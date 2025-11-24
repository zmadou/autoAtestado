from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import re
import os
import datetime
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from queue import Queue
import json
import base64
import ctypes
from ctypes import wintypes

class _DATA_BLOB(ctypes.Structure):
    _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_byte))]

def _protect_data(plain: bytes) -> bytes:
    crypt32 = ctypes.WinDLL("crypt32")
    kernel32 = ctypes.WinDLL("kernel32")

    in_buffer = ctypes.create_string_buffer(plain)
    in_blob = _DATA_BLOB(len(plain), ctypes.cast(in_buffer, ctypes.POINTER(ctypes.c_byte)))
    out_blob = _DATA_BLOB()

    if not crypt32.CryptProtectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob)):
        raise RuntimeError("Falha ao proteger dados (DPAPI)")

    try:
        protected = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return protected
    finally:
        kernel32.LocalFree(out_blob.pbData)

def _unprotect_data(protected: bytes) -> bytes:
    crypt32 = ctypes.WinDLL("crypt32")
    kernel32 = ctypes.WinDLL("kernel32")

    in_buffer = ctypes.create_string_buffer(protected)
    in_blob = _DATA_BLOB(len(protected), ctypes.cast(in_buffer, ctypes.POINTER(ctypes.c_byte)))
    out_blob = _DATA_BLOB()

    if not crypt32.CryptUnprotectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob)):
        raise RuntimeError("Falha ao descriptografar dados (DPAPI)")

    try:
        plain = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return plain
    finally:
        kernel32.LocalFree(out_blob.pbData)

def _cred_path() -> str:
    appdata = os.getenv("APPDATA") or os.path.expanduser("~")
    folder = os.path.join(appdata, "AutoAtestado")
    if not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "credentials.json")

def load_saved_credentials():
    try:
        with open(_cred_path(), "r", encoding="utf-8") as f:
            data = json.load(f)
        user = data.get("username")
        enc = data.get("password")
        if user and enc:
            decrypted = _unprotect_data(base64.b64decode(enc)).decode("utf-8")
            return user, decrypted
    except FileNotFoundError:
        return None, None
    except Exception:
        return None, None
    return None, None

def save_credentials(username: str, password: str):
    try:
        protected = _protect_data(password.encode("utf-8"))
        payload = {"username": username, "password": base64.b64encode(protected).decode("ascii")}
        with open(_cred_path(), "w", encoding="utf-8") as f:
            json.dump(payload, f)
    except Exception:
        pass

def clear_credentials():
    try:
        p = _cred_path()
        if os.path.exists(p):
            os.remove(p)
    except Exception:
        pass

def _settings_path() -> str:
    appdata = os.getenv("APPDATA") or os.path.expanduser("~")
    folder = os.path.join(appdata, "AutoAtestado")
    if not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "settings.json")

def load_settings():
    try:
        with open(_settings_path(), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"attend_reason": "Amparo Legal", "amparo_code": "0000000001", "search_year": "2025"}

def save_settings(settings: dict):
    try:
        with open(_settings_path(), "w", encoding="utf-8") as f:
            json.dump(settings, f)
    except Exception:
        pass
def settings_exists() -> bool:
    try:
        return os.path.exists(_settings_path())
    except Exception:
        return False

class LogManager:
    def __init__(self, log_dir: str = "log"):
        """
        Inicializa o gerenciador de logs com arquivo nomeado por data/hora
        """
        self.log_dir = log_dir
        
        # Criar diretório se não existir
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        # Gerar nome do arquivo com data e hora atual
        now = datetime.datetime.now()
        timestamp = now.strftime("%d_%m_%y__%H_%Mh")
        self.log_file = os.path.join(log_dir, f"log_{timestamp}.txt")
        
        # Criar arquivo de log
        with open(self.log_file, 'w', encoding='utf-8') as f:
            f.write("=== LOG DE LANÇAMENTOS DE AULAS ===\n")
            f.write(f"Arquivo criado em: {now.strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write("-" * 50 + "\n\n")
        
        print(f"📝 Arquivo de log criado: {self.log_file}")
    
    def registrar_lancamento(self, id_aluno: str, aulas_lancadas: list, status: str = "SUCESSO", observacoes: str = ""):
        """
        Registra um lançamento de aulas no log
        """
        timestamp = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        log_entry = f"LANÇAMENTO REGISTRADO\n"
        log_entry += f"Data/Hora: {timestamp}\n"
        log_entry += f"ID do Aluno: {id_aluno}\n"
        log_entry += f"Status: {status}\n"
        log_entry += f"Aulas Processadas: {len(aulas_lancadas)}\n"
        log_entry += f"Detalhes das Aulas: {', '.join(aulas_lancadas) if aulas_lancadas else 'Nenhuma aula processada'}\n"
        
        if observacoes:
            log_entry += f"Observações: {observacoes}\n"
        
        log_entry += "-" * 30 + "\n\n"
        
        try:
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)
            print(f"📝 Log registrado para aluno {id_aluno}")
        except Exception as e:
            print(f"❌ Erro ao registrar no log: {e}")
    
    def registrar_erro(self, id_aluno: str, erro: str):
        """
        Registra um erro no processamento
        """
        timestamp = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        log_entry = f"ERRO REGISTRADO\n"
        log_entry += f"Data/Hora: {timestamp}\n"
        log_entry += f"ID do Aluno: {id_aluno}\n"
        log_entry += f"Erro: {erro}\n"
        log_entry += "-" * 30 + "\n\n"
        
        try:
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)
            print(f"📝 Erro registrado para aluno {id_aluno}")
        except Exception as e:
            print(f"❌ Erro ao registrar erro no log: {e}")


class StopRequested(Exception):
    pass


def processar_atestados(user, password, status_cb=None, resume_event=None, stop_event=None, excel_path='atestados.xlsx', config=None, process_fic=False):
    """
    Executa toda a automação. 
    - status_cb: função para receber mensagens de status (str)
    - resume_event: Event controlado pela UI. Quando limpo (clear), a automação pausa. Quando setado, continua.
    - stop_event: Event para parada total e imediata assim que possível.
    - excel_path: caminho do arquivo da planilha (mantido externo ao programa).
    """
    def notify(msg: str):
        print(msg)
        if status_cb:
            try:
                status_cb(msg)
            except Exception:
                pass

    if resume_event is None:
        resume_event = threading.Event()
        resume_event.set()
    if stop_event is None:
        stop_event = threading.Event()

    def check_abort():
        if stop_event.is_set():
            raise StopRequested("Parado pelo usuário")

    def wait_if_paused():
        # Bloqueia aqui quando em pausa, mas sai se stop for solicitado
        while not resume_event.is_set():
            if stop_event.is_set():
                raise StopRequested("Parado pelo usuário")
            sleep(0.2)

    log_manager = LogManager()
    year = (config.get("search_year") if config else "")
    found_count = 0
    processed_count = 0
    not_found_count = 0

    # Carregar planilha (externa ao programa)
    try:
        excel = load_workbook(excel_path)
    except FileNotFoundError:
        notify(f"Planilha não encontrada: {excel_path}. Deixe 'atestados.xlsx' na mesma pasta do programa.")
        raise
    lista_atestados = excel['Plan1']

    driver = webdriver.Chrome()
    try:
        notify("Abrindo página de login...")
        driver.get('https://senaconline-interno.sp.senac.br/psp/cs90pss/?cmd=login&languageCd=POR')

        # Helpers locais com referência ao driver atual
        def colocar_assim_aparecer(tipo, nome, conteudo, timeout=15):
            try:
                elemento = WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((tipo, nome))
                )
                driver.execute_script("document.activeElement.blur();")
                elemento.clear()
                elemento.send_keys(str(conteudo))
                print("✅ Conteúdo inserido com sucesso.")
            except Exception:
                pass

        def clicar_assim_aparecer(tipo, nome, timeout=15):
            try:
                elemento = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((tipo, nome))
                )
                elemento.click()
                print("✅ Clicou com sucesso.")
            except Exception:
                pass

        check_abort()
        notify("Realizando login...")
        colocar_assim_aparecer(By.ID, 'userid', user)
        colocar_assim_aparecer(By.ID, 'pwd', password)
        clicar_assim_aparecer(By.XPATH, '//*[@id="login"]/div/div[1]/div[8]/input')

        wait_if_paused(); check_abort()
        notify("Acessando lista de atividades por aluno...")
        clicar_assim_aparecer(By.XPATH, """//*[@id="pthnavbca_PORTAL_ROOT_OBJECT"]""")
        clicar_assim_aparecer(By.XPATH, """//*[@id="fldra_HCSR_CURRICULUM_MANAGEMENT"]""")
        clicar_assim_aparecer(By.XPATH, """//*[@id="fldra_HCSR_ATTENDANCE_ROSTER"]""")
        clicar_assim_aparecer(By.XPATH, """//*[@id="crefli_HC_STDNT_ATTENDANCE_GBL"]/a""")

        total_alunos = lista_atestados.max_row - 1
        notify(f"Iniciando processamento de {total_alunos} alunos...")

        # Loop nos IDs da planilha
        for row in lista_atestados.iter_rows(min_row=2, max_row=lista_atestados.max_row, min_col=1, max_col=lista_atestados.max_column):
            wait_if_paused(); check_abort()
            current_id = row[1].value
            current_inicio = row[2].value
            current_fim = row[3].value

            # Se o ID for None ou vazio, significa que chegou ao fim das linhas preenchidas
            if current_id is None or str(current_id).strip() == '':
                notify("Fim das linhas preenchidas alcançado.")
                break

            notify(f"Processando aluno {current_id}...")
            # Lista para armazenar as aulas processadas para este aluno
            aulas_processadas = []
            status_processamento = "SUCESSO"
            observacoes = ""

            try:
                driver.switch_to.default_content()
                WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ptifrmtgtframe")))

                # colocando id
                colocar_assim_aparecer(By.XPATH, """//*[@id="OR_ATND_SRCH_EMPLID"]""", current_id)
                # clicando em pesquisar
                clicar_assim_aparecer(By.XPATH, """//*[@id="#ICSearch"]""")

                driver.switch_to.default_content()
                WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ptifrmtgtframe")))

                # Tabela de resultados
                tabela = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "PTSRCHRESULTS"))
                )

                linhas = driver.find_elements(By.XPATH, '//*[@id="PTSRCHRESULTS"]/tbody/tr')

                encontrou = False
                for linha in linhas:
                    texto = linha.text.upper()
                    if year and year in texto and "EMÉDIO" in texto:
                        link = linha.find_element(By.XPATH, ".//a")
                        link.click()
                        encontrou = True
                        found_count += 1
                        break

                if not encontrou:
                    status_processamento = "ERRO"
                    msg = f"Aluno não encontrado ou sem EMÉDIO {year}" if year else "Aluno não encontrado ou sem EMÉDIO"
                    observacoes = msg
                    log_manager.registrar_lancamento(current_id, aulas_processadas, status_processamento, observacoes)
                    notify(f"Aluno {current_id}: {msg}")
                    not_found_count += 1
                    continue

                driver.switch_to.default_content()
                WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ptifrmtgtframe")))

                # Pega a tabela de aulas
                tabela = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="STDNT_ENRL$scroll$0"]'))
                )

                linhas = tabela.find_elements(By.TAG_NAME, "tr")

                primeira_coluna_elementos = []

                for linha in linhas:
                    colunas = linha.find_elements(By.TAG_NAME, "td")
                    if colunas:
                        primeira_coluna_elementos.append(colunas[0])

                elementos_aulas = []

                for elemento in primeira_coluna_elementos:
                    texto = elemento.text.strip()
                    if re.fullmatch(r'\d+', texto):
                        if len(texto) > 4 and not process_fic:
                            print(f"⚠️ Pulando aula {texto} - mais de 4 dígitos")
                            aulas_processadas.append(f"Aula {texto} - Pulada (mais de 4 dígitos)")
                        else:
                            elementos_aulas.append(elemento)

                notify(f"Encontradas {len(elementos_aulas)} aulas válidas para o aluno {current_id}.")

                # Agora vamos clicar um a um nos elementos de aula
                for i in range(len(elementos_aulas)):
                    wait_if_paused(); check_abort()
                    # Retorna para o contexto padrão
                    driver.switch_to.default_content()
                    WebDriverWait(driver, 20).until(
                        EC.frame_to_be_available_and_switch_to_it((By.ID, "ptifrmtgtframe"))
                    )

                    # Localiza novamente a tabela e os elementos
                    tabela = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="STDNT_ENRL$scroll$0"]'))
                    )
                    linhas = tabela.find_elements(By.TAG_NAME, "tr")

                    primeira_coluna_elementos = []
                    for linha in linhas:
                        colunas = linha.find_elements(By.TAG_NAME, "td")
                        if colunas:
                            primeira_coluna_elementos.append(colunas[0])

                    # Filtra os elementos válidos (apenas números de até 4 dígitos)
                    elementos_aulas_validos = []
                    for elem in primeira_coluna_elementos:
                        texto = elem.text.strip()
                        if re.fullmatch(r'\d+', texto) and (len(texto) <= 4 or process_fic):
                            elementos_aulas_validos.append(elem)

                    # Agora pega o elemento atual baseado no índice do loop
                    elemento = elementos_aulas_validos[i]
                    numero_aula = elemento.text.strip()

                    try:
                        link = elemento.find_element(By.TAG_NAME, "a")
                        driver.execute_script("arguments[0].scrollIntoView();", link)
                        link.click()

                        print(f"✅ Clicou no link do número da aula {i+1}: {numero_aula}")
                        sleep(2)  # Aguarda um pouco após o clique

                    except Exception as e:
                        print(f"❌ Erro ao clicar no link do elemento {i+1}: {e}")
                        continue
                        print(f"❌ Erro ao clicar no link do elemento {i+1}: {e}")
                        continue

                    # Aqui dentro da página aberta, verifica se o texto "Matric" está presente
                    try:
                        tabela_verificacao = WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="ACE_DERIVED_AA2_"]'))
                        )
                        # printar a informação que apararece em tabela_verificacao
                        print(tabela_verificacao.text)

                        if "Matric" in tabela_verificacao.text:
                            # Verificar se também contém "FIC" - se sim, pular
                            if "FIC" in tabela_verificacao.text:
                                print("⚠️ O texto 'Matric' foi encontrado, mas também contém 'FIC'. Pulando esta aula.")
                                clicar_assim_aparecer(By.XPATH, '//*[@id="DERIVED_AA2_DERIVED_LINK10$0"]')
                                # Adicionar aula que foi pulada à lista
                                aulas_processadas.append(f"Aula {numero_aula} - Pulada (Matric + FIC)")
                                sleep(2)
                                continue

                            print("✅ O texto 'Matric' foi encontrado na tabela (sem FIC).")
                            clicar_assim_aparecer(By.XPATH, '//*[@id="ICTAB_1"]')

                            # Inserir as datas e informações
                            colocar_assim_aparecer(By.XPATH, '//*[@id="DIG_APR_EST_WRK_START_DT"]', current_inicio.strftime('%d/%m/%Y'))
                            colocar_assim_aparecer(By.XPATH, '//*[@id="DIG_APR_EST_WRK_END_DT"]', current_fim.strftime('%d/%m/%Y'))

                            select_element = WebDriverWait(driver, 15).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="DIG_APR_EST_WRK_ATTEND_REASON"]'))
                            )
                            selected_reason = (config.get("attend_reason") if config else "Amparo Legal")
                            for option in select_element.find_elements(By.TAG_NAME, 'option'):
                                if option.text.strip() == selected_reason:
                                    option.click()
                                    print(f"✅ Opção '{selected_reason}' selecionada.")
                                    break
                            if selected_reason == "Amparo Legal":
                                selected_code = (config.get("amparo_code") if config else "0000000001")
                                colocar_assim_aparecer(By.XPATH, '//*[@id="DIG_APR_EST_WRK_REASON_DESCR"]', selected_code)

                            print("✅ Dados inseridos com sucesso.")

                            # clicar no botao de aplicar
                            clicar_assim_aparecer(By.XPATH, '//*[@id="DIG_APR_EST_WRK_PROCESS_BTN"]')
                            # clicar em botao de salvar
                            clicar_assim_aparecer(By.XPATH, '//*[@id="#ICSave"]')
                            # clicar no link de voltar para a lista
                            clicar_assim_aparecer(By.XPATH, '//*[@id="DERIVED_AA2_DERIVED_LINK10$0"]')

                            # Adicionar aula processada com sucesso à lista
                            aulas_processadas.append(f"Aula {numero_aula} - Lançamento realizado")
                            processed_count += 1
                            sleep(2)
                            
                        else:
                            print("❌ O texto 'Matric' NÃO foi encontrado na tabela.")
                            clicar_assim_aparecer(By.XPATH, '//*[@id="DERIVED_AA2_DERIVED_LINK10$0"]')
                            # Adicionar aula que não foi processada à lista
                            aulas_processadas.append(f"Aula {numero_aula} - Não processada (sem Matric)")
                            sleep(2)

                    except Exception as e:
                        print(f"❌ Erro ao buscar a tabela: {e}")
                        aulas_processadas.append(f"Aula {numero_aula} - Erro: {str(e)}")

                    # Se necessário, clique para retornar após cada verificação, para não quebrar o próximo loop
                    try:
                        voltar = driver.find_element(By.XPATH, '//*[@id="DERIVED_AA2_DERIVED_LINK10$0"]')
                        voltar.click()
                        sleep(2)
                    except Exception as e:
                        print(f"❌ Erro ao retornar: {e}")

            except StopRequested:
                notify("Parada solicitada. Encerrando processamento atual...")
                break
            except Exception as e:
                status_processamento = "ERRO"
                observacoes = f"Erro geral no processamento: {str(e)}"
                log_manager.registrar_erro(current_id, str(e))

            # Registrar o lançamento no log para este aluno
            if aulas_processadas:
                observacoes_final = f"Período: {current_inicio.strftime('%d/%m/%Y')} a {current_fim.strftime('%d/%m/%Y')}"
                if observacoes:
                    observacoes_final += f" | {observacoes}"
                log_manager.registrar_lancamento(current_id, aulas_processadas, status_processamento, observacoes_final)
            else:
                log_manager.registrar_lancamento(current_id, [], "SEM_AULAS", "Nenhuma aula foi encontrada para processar")

            check_abort()
            driver.switch_to.default_content()
            # Acessando lista de atividades por aluno
            clicar_assim_aparecer(By.XPATH, """//*[@id=\"pthnavbca_PORTAL_ROOT_OBJECT\"]""")
            clicar_assim_aparecer(By.XPATH, """//*[@id=\"fldra_HCSR_CURRICULUM_MANAGEMENT\"]""")
            clicar_assim_aparecer(By.XPATH, """//*[@id=\"fldra_HCSR_ATTENDANCE_ROSTER\"]""")
            clicar_assim_aparecer(By.XPATH, """//*[@id=\"crefli_HC_STDNT_ATTENDANCE_GBL\"]/a""")

            notify(f"Aluno {current_id} finalizado.")

        if found_count == 0:
            if year:
                notify(f"Nenhuma linha encontrada para EMÉDIO {year}.")
            else:
                notify("Nenhuma linha encontrada.")
        else:
            notify("🎉 Processamento concluído! Verifique o arquivo de log na pasta 'log'.")
        return {"found_count": found_count, "processed_count": processed_count, "not_found_count": not_found_count, "year": year}

    except StopRequested:
        notify("Execução interrompida pelo usuário.")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("AutoAtestado - SENAC")
        self.geometry("720x580")
        self.resizable(False, False)

        # Adicionar ícone se existir
        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass  # Se não encontrar o ícone, continua sem ele

        # Estado
        self.worker_thread = None
        self.resume_event = threading.Event(); self.resume_event.set()
        self.stop_event = threading.Event()
        self.is_running = False
        self.is_paused = False
        self.pending_restart = False
        self.status_queue = Queue()
        self.settings_open = False
        self.status_order = []
        self.status_by_id = {}
        self.planilha_preview = []
        self.completed_order = []
        self.active_order = []
        self.error_order = []

        # Estado
        self.worker_thread = None
        self.resume_event = threading.Event(); self.resume_event.set()
        self.stop_event = threading.Event()
        self.is_running = False
        self.is_paused = False
        self.processo_concluido = False
        self.processo_erro = None

        # UI
        padding = {"padx": 8, "pady": 6}
        frm = ttk.Frame(self)
        frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=10) 
        try:
            for i in range(4):
                frm.grid_columnconfigure(i, weight=1)
            frm.grid_columnconfigure(0, weight=0)
            frm.grid_columnconfigure(1, weight=1)
            frm.grid_columnconfigure(2, weight=1)
            frm.grid_columnconfigure(3, weight=1)
        except Exception:
            pass

        hdr = ttk.Frame(frm)
        hdr.grid(row=0, column=0, columnspan=4, sticky="ew")
        try:
            hdr.grid_columnconfigure(0, weight=1)
            hdr.grid_columnconfigure(1, weight=0)
        except Exception:
            pass
        self.settings_btn = ttk.Button(hdr, text="Config", width=8, command=self.open_settings)
        self.settings_btn.grid(row=0, column=1, sticky="e", padx=4, pady=2)

        ttk.Label(frm, text="Usuário:").grid(row=1, column=0, sticky="w", **padding)
        self.user_var = tk.StringVar()
        self.user_entry = ttk.Entry(frm, textvariable=self.user_var, width=28)
        self.user_entry.grid(row=1, column=1, columnspan=2, sticky="w", **padding)

        ttk.Label(frm, text="Senha:").grid(row=2, column=0, sticky="w", **padding)
        self.pass_var = tk.StringVar()
        self.pass_entry = ttk.Entry(frm, textvariable=self.pass_var, width=28, show="*")
        self.pass_entry.grid(row=2, column=1, columnspan=2, sticky="w", **padding)

        self.remember_var = tk.BooleanVar(value=False)
        self.remember_cb = ttk.Checkbutton(frm, text="Lembrar credenciais", variable=self.remember_var)
        self.remember_cb.grid(row=3, column=0, columnspan=4, sticky="w", **padding)
        self.process_fic_var = tk.BooleanVar(value=False)
        self.process_fic_cb = ttk.Checkbutton(frm, text="Lançar amparo em aulas FIC", variable=self.process_fic_var)
        self.process_fic_cb.grid(row=4, column=0, columnspan=4, sticky="w", **padding)

        btns = ttk.Frame(frm)
        btns.grid(row=5, column=0, columnspan=4, sticky="ew")
        try:
            btns.grid_columnconfigure(0, weight=1)
            btns.grid_columnconfigure(1, weight=1)
        except Exception:
            pass
        self.start_btn = ttk.Button(btns, text="Iniciar", command=self.on_start)
        self.start_btn.grid(row=0, column=0, sticky="ew", **padding)
        self.pause_btn = ttk.Button(btns, text="Pausar", command=self.on_pause_resume, state=tk.DISABLED)
        self.pause_btn.grid(row=0, column=1, sticky="ew", **padding)
        self.stop_btn = ttk.Button(btns, text="Parar", command=self.on_stop, state=tk.DISABLED)
        self.stop_btn.grid(row=1, column=0, sticky="ew", **padding)
        self.restart_btn = ttk.Button(btns, text="Reiniciar", command=self.on_restart)
        self.restart_btn.grid(row=1, column=1, sticky="ew", **padding)

        ttk.Separator(frm).grid(row=6, column=0, columnspan=4, sticky="ew", pady=(10, 0))

        ttk.Label(frm, text="Status:").grid(row=7, column=0, columnspan=4, sticky="w", **padding)
        self.status_var = tk.StringVar(value="Pronto.")
        self.status_lbl = ttk.Label(frm, textvariable=self.status_var, wraplength=400, anchor="w", justify="left")
        self.status_lbl.grid(row=7, column=0, columnspan=4, sticky="w", **padding)
        self.progress_text = tk.StringVar(value="Progresso: 0/0")
        ttk.Label(frm, textvariable=self.progress_text).grid(row=8, column=0, columnspan=4, sticky="w", padx=10)
        self.progress = ttk.Progressbar(frm, mode="determinate")
        self.progress.grid(row=9, column=0, columnspan=4, sticky="ew", padx=10)
        self.status_tree = ttk.Treeview(frm, columns=("id", "periodo", "status"), show="headings", height=10)
        self.status_tree.heading("id", text="ID")
        self.status_tree.heading("periodo", text="Período")
        self.status_tree.heading("status", text="Status")
        self.status_tree.column("id", width=140, anchor="w")
        self.status_tree.column("periodo", width=200, anchor="w")
        self.status_tree.column("status", width=100, anchor="w")
        self.status_tree.grid(row=10, column=0, columnspan=3, sticky="nsew", padx=10)
        self.status_scroll = ttk.Scrollbar(frm, orient="vertical", command=self.status_tree.yview)
        self.status_scroll.grid(row=10, column=3, sticky="ns")
        self.status_tree.configure(yscrollcommand=self.status_scroll.set)
        try:
            self.status_tree.tag_configure("concluído", foreground="#2e7d32")
            self.status_tree.tag_configure("carregando", foreground="#f9a825")
            self.status_tree.tag_configure("erro", foreground="#c62828")
            self.status_tree.tag_configure("aguardando", foreground="#616161")
            frm.grid_rowconfigure(10, weight=1)
        except Exception:
            pass

        # Info da planilha (externa)
        ttk.Label(frm, text="Planilha: 'atestados.xlsx' (mesma pasta do programa)").grid(row=11, column=0, columnspan=4, sticky="w", padx=8)

        self.bind('<Return>', lambda e: self.on_start() if (not self.is_running and not self.settings_open) else None)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # Poll da fila de status
        self.after(200, self._poll_status)

        # Carregar credenciais salvas, se existirem
        try:
            u, p = load_saved_credentials()
            if u and p:
                self.user_var.set(u)
                self.pass_var.set(p)
                self.remember_var.set(True)
        except Exception:
            pass
        s = load_settings()
        self.attend_reason = s.get("attend_reason", "Amparo Legal")
        self.amparo_code = s.get("amparo_code", "0000000001")
        self.search_year = s.get("search_year", "2025")
        try:
            wb = load_workbook('atestados.xlsx')
            ws = wb['Plan1']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                idv = row[1].value
                ini = row[2].value
                fim = row[3].value
                if idv and str(idv).strip() != '':
                    id_s = str(idv)
                    self.planilha_preview.append((id_s, ini, fim))
                    self.status_by_id[id_s] = "aguardando"
                    self.status_order.append(id_s)
            total = len(self.planilha_preview)
            self.progress.configure(mode="determinate", maximum=max(1, total))
            self.progress['value'] = 0
            self.progress_text.set(f"Progresso: 0/{total}")
            self._render_status()
        except Exception:
            pass

    def _enqueue_status(self, msg: str):
        self.status_queue.put(msg)

    def _poll_status(self):
        try:
            while True:
                msg = self.status_queue.get_nowait()
                m = re.search(r'Processando aluno\s+(\d+)', msg)
                if m:
                    self.status_by_id[m.group(1)] = "carregando"
                    if m.group(1) not in self.active_order:
                        self.active_order.append(m.group(1))
                    if m.group(1) in self.completed_order:
                        self.completed_order.remove(m.group(1))
                    if m.group(1) in self.error_order:
                        self.error_order.remove(m.group(1))
                    self._render_status()
                    continue
                m = re.search(r'Aluno\s+(\d+)\s+finalizado\.', msg)
                if m:
                    self.status_by_id[m.group(1)] = "concluído"
                    if m.group(1) in self.active_order:
                        self.active_order.remove(m.group(1))
                    if m.group(1) not in self.completed_order:
                        self.completed_order.append(m.group(1))
                    self._render_status()
                    continue
                m = re.search(r'Aluno\s+(\d+):', msg)
                if m:
                    self.status_by_id[m.group(1)] = "erro"
                    if m.group(1) in self.active_order:
                        self.active_order.remove(m.group(1))
                    if m.group(1) not in self.error_order:
                        self.error_order.append(m.group(1))
                    self._render_status()
                    continue
                self.status_var.set(msg)
        except Exception:
            pass
        # Se houver reinício pendente e não estiver rodando, reinicia
        if self.pending_restart and not self.is_running:
            self.pending_restart = False
            self.on_start()
        self.after(200, self._poll_status)

    def _render_status(self):
        lines = []
        def fmt(idv, ini, fim):
            st = self.status_by_id.get(idv, "aguardando")
            d1 = ini.strftime('%d/%m/%Y') if hasattr(ini, 'strftime') else (ini if ini else '')
            d2 = fim.strftime('%d/%m/%Y') if hasattr(fim, 'strftime') else (fim if fim else '')
            return f"{idv} - {d1} a {d2} [{st}]"

        preview_map = {idv: (ini, fim) for idv, ini, fim in self.planilha_preview}

        for idv in self.completed_order:
            if idv in preview_map:
                ini, fim = preview_map[idv]
                lines.append(fmt(idv, ini, fim))
        for idv in self.error_order:
            if idv in preview_map and idv not in self.completed_order:
                ini, fim = preview_map[idv]
                lines.append(fmt(idv, ini, fim))
        for idv, ini, fim in self.planilha_preview:
            if idv not in self.completed_order and idv not in self.error_order and idv not in self.active_order:
                lines.append(fmt(idv, ini, fim))
        for idv in self.active_order:
            if idv in preview_map:
                ini, fim = preview_map[idv]
                lines.append(fmt(idv, ini, fim))
        try:
            for item in self.status_tree.get_children():
                self.status_tree.delete(item)
            def insert_line(idv, ini, fim):
                st = self.status_by_id.get(idv, "aguardando")
                d1 = ini.strftime('%d/%m/%Y') if hasattr(ini, 'strftime') else (ini if ini else '')
                d2 = fim.strftime('%d/%m/%Y') if hasattr(fim, 'strftime') else (fim if fim else '')
                self.status_tree.insert('', 'end', values=(idv, f"{d1} a {d2}", st), tags=(st,))
            for idv in self.completed_order:
                if idv in preview_map:
                    ini, fim = preview_map[idv]
                    insert_line(idv, ini, fim)
            for idv in self.error_order:
                if idv in preview_map and idv not in self.completed_order:
                    ini, fim = preview_map[idv]
                    insert_line(idv, ini, fim)
            for idv, ini, fim in self.planilha_preview:
                if idv not in self.completed_order and idv not in self.error_order and idv not in self.active_order:
                    insert_line(idv, ini, fim)
            for idv in self.active_order:
                if idv in preview_map:
                    ini, fim = preview_map[idv]
                    insert_line(idv, ini, fim)
            total = len(self.planilha_preview)
            done = len(self.completed_order)
            try:
                self.progress.configure(mode="determinate", maximum=max(1, total))
                self.progress['value'] = done
            except Exception:
                pass
            try:
                self.progress_text.set(f"Progresso: {done}/{total}")
            except Exception:
                pass
        except Exception:
            self.status_var.set("\n".join(lines) if lines else "Pronto.")

    def _position_right(self):
        try:
            self.update_idletasks()
            w = self.winfo_width() or 460
            h = self.winfo_height() or 300
            sw = self.winfo_screenwidth()
            x = max(0, sw - w - 20)
            y = 20
            self.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass


    def on_start(self):
        if self.is_running:
            return
        user = self.user_var.get().strip()
        pwd = self.pass_var.get().strip()
        if not user or not pwd:
            messagebox.showwarning("Dados obrigatórios", "Informe usuário e senha.")
            return
        allowed_reasons = ["Amparo Legal", "Aproveitamento de Estudos", "Matrícula Fora do Prazo"]
        if not settings_exists() or not self.search_year or not self.attend_reason or self.attend_reason not in allowed_reasons or (self.attend_reason == "Amparo Legal" and not self.amparo_code):
            messagebox.showwarning("Configurações necessárias", "Antes de iniciar, abra Config, escolha Ano e Razão e clique em Salvar.")
            return
        if self.remember_var.get():
            save_credentials(user, pwd)
        else:
            clear_credentials()
        self.is_running = True
        self.is_paused = False
        self.resume_event.set()
        self.stop_event.clear()

        # Travar inputs enquanto roda
        self.user_entry.configure(state=tk.DISABLED)
        self.pass_entry.configure(state=tk.DISABLED)
        self.start_btn.configure(state=tk.DISABLED)
        self.pause_btn.configure(state=tk.NORMAL, text="Pausar")
        self.stop_btn.configure(state=tk.NORMAL)

        self.status_var.set("Iniciando automação...")
        try:
            self._position_right()
        except Exception:
            pass
        try:
            total = len(self.planilha_preview)
            self.progress.configure(mode="determinate", maximum=max(1, total))
            self.progress['value'] = len(self.completed_order)
            self.progress_text.set(f"Progresso: {len(self.completed_order)}/{total}")
        except Exception:
            pass

        def target():
            try:
                cfg = {"attend_reason": self.attend_reason, "amparo_code": self.amparo_code, "search_year": self.search_year}
                result = processar_atestados(user, pwd, status_cb=self._enqueue_status, resume_event=self.resume_event, stop_event=self.stop_event, config=cfg, process_fic=self.process_fic_var.get())
                if not self.stop_event.is_set():
                    if result and result.get("found_count", 0) == 0:
                        yr = result.get("year") or ""
                        msg = f"Nenhuma linha encontrada para EMÉDIO {yr}." if yr else "Nenhuma linha encontrada."
                        self._enqueue_status(msg)
                        self.processo_erro = msg
                    else:
                        self._enqueue_status("🎉 Processamento concluído! Verifique o arquivo de log na pasta 'log'.")
                        self.processo_concluido = True
            except Exception as e:
                self._enqueue_status(f"❌ Erro: {e}")
                self.processo_erro = str(e)
            finally:
                def restore():
                    self.is_running = False
                    self.is_paused = False
                    self.user_entry.configure(state=tk.NORMAL)
                    self.pass_entry.configure(state=tk.NORMAL)
                    self.start_btn.configure(state=tk.NORMAL)
                    self.pause_btn.configure(state=tk.DISABLED, text="Pausar")
                    self.stop_btn.configure(state=tk.DISABLED)
                    self.restart_btn.configure(state=tk.DISABLED)
                    try:
                        self.progress.stop()
                        self.progress.configure(mode="determinate")
                    except Exception:
                        pass

                    # Verificar se deve mostrar pop-up de conclusão
                    if hasattr(self, 'processo_concluido') and self.processo_concluido:
                        messagebox.showinfo("AutoAtestado - Concluído", "Processamento concluído com sucesso!\n\nVerifique o arquivo de log na pasta 'log' para detalhes.")
                        self.processo_concluido = False
                    elif hasattr(self, 'processo_erro') and self.processo_erro:
                        messagebox.showerror("AutoAtestado - Erro", f"Falha na execução:\n\n{self.processo_erro}")
                        self.processo_erro = None
                self.after(0, restore)

        self.worker_thread = threading.Thread(target=target, daemon=True)
        self.worker_thread.start()

    def open_settings(self):
        dialog = tk.Toplevel(self)
        dialog.title("Configurações")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()
        self.settings_open = True
        try:
            self.start_btn.configure(state=tk.DISABLED)
        except Exception:
            pass
        pad = {"padx": 10, "pady": 8}
        frm = ttk.Frame(dialog)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text="Ano (4 dígitos):").grid(row=0, column=0, sticky="w", **pad)
        year_var = tk.StringVar(value=str(self.search_year))
        year_entry = ttk.Entry(frm, textvariable=year_var, width=10)
        def validate_year(P):
            return P.isdigit() and len(P) <= 4
        vcmd = (self.register(validate_year), '%P')
        year_entry.configure(validate='key', validatecommand=vcmd)
        year_entry.grid(row=0, column=1, sticky="w", **pad)
        ttk.Label(frm, text="Razão de ausência:").grid(row=1, column=0, sticky="w", **pad)
        reasons = ["Amparo Legal", "Aproveitamento de Estudos", "Matrícula Fora do Prazo"]
        reason_var = tk.StringVar(value=self.attend_reason)
        reason_cb = ttk.Combobox(frm, textvariable=reason_var, values=reasons, state="readonly", width=28)
        reason_cb.grid(row=1, column=1, sticky="w", **pad)
        ttk.Label(frm, text="Motivo (Amparo Legal):").grid(row=2, column=0, sticky="w", **pad)
        amparo_options = [
            ("0000000001", "Problemas de saúde"),
            ("0000000002", "Licença Maternidade"),
            ("0000000003", "Adoção"),
            ("0000000004", "Licença paternidade"),
            ("0000000005", "Serviço militar"),
            ("0000000006", "Representação desportiva"),
            ("0000000008", "Educação Física"),
            ("0000000009", "Crença religiosa"),
        ]
        amparo_display = [f"{code} - {desc}" for code, desc in amparo_options]
        amparo_var = tk.StringVar()
        try:
            current_display = next(f"{c} - {d}" for c, d in amparo_options if c == self.amparo_code)
            amparo_var.set(current_display)
        except Exception:
            amparo_var.set(amparo_display[0])
        amparo_cb = ttk.Combobox(frm, textvariable=amparo_var, values=amparo_display, state="readonly", width=28)
        amparo_cb.grid(row=2, column=1, sticky="w", **pad)
        def on_reason_change(*_):
            if reason_var.get() == "Amparo Legal":
                amparo_cb.configure(state="readonly")
            else:
                amparo_cb.configure(state="disabled")
        reason_var.trace_add("write", on_reason_change)
        on_reason_change()
        btns = ttk.Frame(frm)
        btns.grid(row=3, column=0, columnspan=2, sticky="e", **pad)
        def on_save():
            self.attend_reason = reason_var.get()
            if self.attend_reason == "Amparo Legal":
                sel = amparo_var.get()
                code = sel.split(" - ")[0]
                self.amparo_code = code
            yr = year_var.get()
            yr_digits = "".join(ch for ch in yr if ch.isdigit())[:4]
            if yr_digits:
                self.search_year = yr_digits
            s = {"attend_reason": self.attend_reason, "amparo_code": self.amparo_code}
            s["search_year"] = self.search_year
            save_settings(s)
            dialog.destroy()
            self.settings_open = False
            try:
                self.start_btn.configure(state=tk.NORMAL)
            except Exception:
                pass
        ttk.Button(btns, text="Salvar", command=on_save).pack(side=tk.RIGHT, padx=4)
        def on_cancel():
            dialog.destroy()
            self.settings_open = False
            try:
                self.start_btn.configure(state=tk.NORMAL)
            except Exception:
                pass
        ttk.Button(btns, text="Cancelar", command=on_cancel).pack(side=tk.RIGHT)

    def on_pause_resume(self):
        if not self.is_running:
            return
        if not self.is_paused:
            # Pausar
            self.resume_event.clear()
            self.is_paused = True
            self.pause_btn.configure(text="Retomar")
            self.status_var.set("Pausado. Clique em Retomar para continuar.")
        else:
            # Retomar
            self.resume_event.set()
            self.is_paused = False
            self.pause_btn.configure(text="Pausar")
            self.status_var.set("Retomando...")

    def on_stop(self):
        if not self.is_running:
            return
        if messagebox.askyesno("Parar", "Deseja interromper o processamento atual?"):
            self.stop_event.set()
            self.resume_event.set()  # garante sair da pausa
            self.status_var.set("Parando... aguarde a etapa atual finalizar.")
            # Botões: desabilita parar para evitar múltiplos cliques
            self.stop_btn.configure(state=tk.DISABLED)

    def on_restart(self):
        if self.is_running:
            if not messagebox.askyesno("Reiniciar", "Isso vai interromper a execução atual e iniciar do começo da planilha. Continuar?"):
                return
            self.pending_restart = True
            self.on_stop()
        else:
            self.on_start()

    def on_close(self):
        if self.is_running:
            if not messagebox.askyesno("Sair", "A automação ainda está em execução. Deseja encerrar mesmo assim?"):
                return
            # Solicita parada
            self.stop_event.set()
            self.resume_event.set()
        try:
            pass
        finally:
            self.destroy()


if __name__ == '__main__':
    app = App()
    app.mainloop()
